import os
from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from telegram import Update
from telegram.ext import ContextTypes

from src.config import settings
from src.database import db
from src.logger import logger
from src.models.schemas import FinanceCategory


async def gasto_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    message = update.effective_message
    user = update.effective_user
    if not message or not user:
        return

    args = context.args
    if not args or len(args) < 2:
        await message.reply_text(
            "Usa: /gasto <monto> <categoría> [descripción]\n"
            "Ejemplo: `/gasto 150.50 despensa Compras del supermercado`\n\n"
            "Categorías sugeridas: alimentacion, transporte, servicios, "
            "entretenimiento, salud, educacion, otros"
        )
        return

    amount_str = args[0].replace(",", ".")
    try:
        amount = float(amount_str)
        if amount <= 0:
            raise ValueError
    except ValueError:
        await message.reply_text("El monto debe ser un número positivo. Ej: 150.50")
        return

    category = args[1].lower()
    description = " ".join(args[2:]) if len(args) > 2 else None

    record_id = await db.add_finance_record(
        chat_id=user.id,
        amount=amount,
        category=FinanceCategory.expense.value,
        subcategory=category,
        description=description,
        currency=settings.default_currency,
    )

    logger.info(
        "Expense recorded: user=%d amount=%.2f category=%s id=%d",
        user.id, amount, category, record_id,
    )

    await message.reply_text(
        f"✅ Gasto registrado:\n"
        f"   • Monto: {amount:,.2f} {settings.default_currency}\n"
        f"   • Categoría: {category}\n"
        + (f"   • Descripción: {description}\n" if description else "")
        + f"   • ID: {record_id}"
    )


async def ingreso_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    message = update.effective_message
    user = update.effective_user
    if not message or not user:
        return

    args = context.args
    if not args or len(args) < 2:
        await message.reply_text(
            "Usa: /ingreso <monto> <fuente> [descripción]\n"
            "Ejemplo: `/ingreso 15000 nomina Sueldo mensual`\n\n"
            "Fuentes sugeridas: nomina, freelance, inversiones, ventas, otros"
        )
        return

    amount_str = args[0].replace(",", ".")
    try:
        amount = float(amount_str)
        if amount <= 0:
            raise ValueError
    except ValueError:
        await message.reply_text("El monto debe ser un número positivo. Ej: 15000")
        return

    source = args[1].lower()
    description = " ".join(args[2:]) if len(args) > 2 else None

    record_id = await db.add_finance_record(
        chat_id=user.id,
        amount=amount,
        category=FinanceCategory.income.value,
        subcategory=source,
        description=description,
        currency=settings.default_currency,
    )

    logger.info(
        "Income recorded: user=%d amount=%.2f source=%s id=%d",
        user.id, amount, source, record_id,
    )

    await message.reply_text(
        f"✅ Ingreso registrado:\n"
        f"   • Monto: {amount:,.2f} {settings.default_currency}\n"
        f"   • Fuente: {source}\n"
        + (f"   • Descripción: {description}\n" if description else "")
        + f"   • ID: {record_id}"
    )


async def finanzas_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    message = update.effective_message
    user = update.effective_user
    if not message or not user:
        return

    await message.reply_chat_action("typing")

    now = datetime.utcnow()
    start_date = now.replace(day=1).strftime("%Y-%m-%d 00:00:00")
    end_date = now.strftime("%Y-%m-%d 23:59:59")

    summary = await db.get_finance_summary(
        chat_id=user.id, start_date=start_date, end_date=end_date
    )

    if summary["transaction_count"] == 0:
        await message.reply_text(
            "No hay registros financieros este mes. "
            "Usa /gasto o /ingreso para comenzar."
        )
        return

    lines = [
        f"📊 *Resumen Financiero - {now.strftime('%B %Y')}*",
        "",
        f"💵 *Ingresos:* {summary['total_income']:,.2f} {settings.default_currency}",
        f"💸 *Gastos:* {summary['total_expenses']:,.2f} {settings.default_currency}",
        f"💰 *Balance:* {summary['balance']:,.2f} {settings.default_currency}",
        "",
    ]

    if summary["expense_by_category"]:
        lines.append("*Gastos por categoría:*")
        for cat, total in sorted(
            summary["expense_by_category"].items(),
            key=lambda x: x[1],
            reverse=True,
        ):
            lines.append(f"   • {cat}: {total:,.2f} {settings.default_currency}")
        lines.append("")

    if summary["income_by_category"]:
        lines.append("*Ingresos por fuente:*")
        for cat, total in sorted(
            summary["income_by_category"].items(),
            key=lambda x: x[1],
            reverse=True,
        ):
            lines.append(f"   • {cat}: {total:,.2f} {settings.default_currency}")

    lines.append(f"\n📝 Total de transacciones: {summary['transaction_count']}")

    await message.reply_text("\n".join(lines), parse_mode="Markdown")


async def exportar_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    message = update.effective_message
    user = update.effective_user
    if not message or not user:
        return

    args = context.args
    export_type = args[0].lower() if args else "finanzas"

    await message.reply_chat_action("typing")

    if export_type in ("finanzas", "finanzas_completo"):
        await _export_finances(update, user.id, export_type == "finanzas_completo")
    else:
        await message.reply_text(
            "Tipos de exportación disponibles:\n"
            "• `/exportar finanzas` - Resumen del mes actual\n"
            "• `/exportar finanzas_completo` - Todos los registros"
        )


async def _export_finances(
    update: Update, chat_id: int, full_history: bool = False
) -> None:
    message = update.effective_message
    if not message:
        return

    export_id = await db.create_export(chat_id, "finanzas")

    try:
        settings.obsidian_finanzas_path.mkdir(parents=True, exist_ok=True)
        filename = f"finanzas_{chat_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = settings.obsidian_finanzas_path / filename

        if full_history:
            records = await db.get_finance_records(chat_id)
        else:
            now = datetime.utcnow()
            start = now.replace(day=1).strftime("%Y-%m-%d 00:00:00")
            end = now.strftime("%Y-%m-%d 23:59:59")
            records = await db.get_finance_records(chat_id, start, end)

        if not records:
            await db.update_export(export_id, "failed", error_message="No records found")
            await message.reply_text("No hay registros financieros para exportar.")
            return

        df = pd.DataFrame(records)

        columns_map = {
            "id": "ID",
            "amount": "Monto",
            "category": "Categoría",
            "subcategory": "Subcategoría",
            "description": "Descripción",
            "currency": "Moneda",
            "recorded_at": "Fecha",
        }
        df = df.rename(columns=columns_map)
        df = df[[c for c in columns_map.values() if c in df.columns]]

        with pd.ExcelWriter(str(filepath), engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Finanzas", index=False)
            worksheet = writer.sheets["Finanzas"]

            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font_white = Font(bold=True, size=11, color="FFFFFF")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for col_idx, _ in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border

            for col_idx, col_name in enumerate(df.columns, 1):
                max_len = max(
                    df[col_name].astype(str).map(len).max() if len(df) > 0 else 0,
                    len(col_name),
                )
                worksheet.column_dimensions[chr(64 + col_idx)].width = min(max_len + 4, 50)

        summary = await db.get_finance_summary(chat_id)
        summary_df = pd.DataFrame([{
            "Concepto": "Ingresos",
            "Monto": summary["total_income"],
        }, {
            "Concepto": "Gastos",
            "Monto": summary["total_expenses"],
        }, {
            "Concepto": "Balance",
            "Monto": summary["balance"],
        }])

        with pd.ExcelWriter(str(filepath), engine="openpyxl", mode="a") as writer:
            summary_df.to_excel(writer, sheet_name="Resumen", index=False)

        await db.update_export(export_id, "completed", file_path=str(filepath))

        file_size = os.path.getsize(str(filepath))
        size_str = (
            f"{file_size / 1024:.1f} KB"
            if file_size < 1024 * 1024
            else f"{file_size / (1024 * 1024):.1f} MB"
        )

        await message.reply_text(
            f"✅ Exportación completada:\n"
            f"   • Archivo: {filename}\n"
            f"   • Tamaño: {size_str}\n"
            f"   • Registros: {len(records)}\n"
            f"   • Ruta local: {filepath}"
        )

        logger.info(
            "Export completed: user=%d file=%s records=%d",
            chat_id, filename, len(records),
        )

    except Exception as e:
        logger.exception("Export failed for user %d", chat_id)
        await db.update_export(export_id, "failed", error_message=str(e))
        await message.reply_text(f"❌ Error al exportar: {e}")
